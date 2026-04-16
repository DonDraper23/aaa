#!/usr/bin/env python3
"""
Consignee OCR 自动识别脚本
引擎：百度 AI Studio Layout Parsing API
验证：字段格式硬校验（邮箱/电话/护照号格式规则）
输出：Excel（已验证绿色 / 需人工核查红色）

依赖安装：
    pip install pdf2image openpyxl requests Pillow
    brew install poppler
"""

import base64
import io
import re
import time
import requests
import openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ============================================================
# ★ 配置区
# ============================================================

API_URL = "https://c37133u2wdhdgbf3.aistudio-app.com/layout-parsing"
TOKEN   = "656cbc2d6ca0150e2aad03d659a112e27d9aaa50"

PDF_FOLDER   = "/Users/GHtextile/Desktop/2026order/3-3月份大订单/01-客户资料/NAD consignee"
OUTPUT_EXCEL = "/Users/GHtextile/Desktop/consignee_output.xlsx"

# PDF 转图片分辨率（150 dpi 已够用，200 dpi 更清晰但更慢）
DPI = 150

# 已知车型关键词（用于验证车型字段）
KNOWN_CARS = ["coolray", "t-roc", "troc", "tiguan", "polo", "golf",
              "passat", "kodiaq", "tucson", "audi"]

# 已知颜色关键词
KNOWN_COLORS = ["white", "black", "grey", "gray", "red", "blue",
                "silver", "brown", "beige", "green"]

# ============================================================
# OCR 调用
# ============================================================

def _post_image(img, quality: int = 82, timeout: int = 90, retries: int = 3) -> list[dict]:
    """将 PIL Image 发送到百度 API，返回排序后的 blocks，失败自动重试"""
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=quality)
    img_b64 = base64.b64encode(buf.getvalue()).decode()

    for attempt in range(1, retries + 1):
        try:
            resp = requests.post(
                API_URL,
                headers={"Authorization": f"token {TOKEN}", "Content-Type": "application/json"},
                json={"file": img_b64},
                timeout=timeout,
            )
            resp.raise_for_status()
            blocks = (resp.json()
                      ["result"]["layoutParsingResults"][0]
                      ["prunedResult"]["parsing_res_list"])
            return sorted(
                [b for b in blocks if b.get("block_content", "").strip()],
                key=lambda b: b["block_bbox"][1]
            )
        except Exception as e:
            if attempt == retries:
                raise
            wait = attempt * 5  # 5s, 10s, 15s
            print(f"    ⚠️ 第{attempt}次超时，{wait}s后重试...")
            time.sleep(wait)


def call_baidu_api(pdf_path: str) -> tuple[list[dict], list[dict]]:
    """
    两次调用：
      1. 全图（150 dpi）→ 获取联系信息、车型、以及大部分护照数据
      2. 裁剪上半（0-55%，120 dpi）→ 补充识别护照号/姓名（处理全图漏识别的情况）
    返回 (full_blocks, top_blocks)
    """
    from pdf2image import convert_from_path

    img_full = convert_from_path(pdf_path, dpi=DPI)[0]

    # 第一次：全图
    full_blocks = _post_image(img_full, quality=82)

    # 第二次：裁剪上半（护照区域），用于补充护照号/姓名
    w, h = img_full.size
    img_small = img_full.resize((int(w * 0.75), int(h * 0.75)))  # 缩小后再裁
    crop = img_small.crop((0, 0, img_small.width, int(img_small.height * 0.58)))
    top_blocks = _post_image(crop, quality=75)

    return full_blocks, top_blocks


# ============================================================
# 字段提取
# ============================================================

_RE_EMAIL    = re.compile(r"[\w.\-+]+@[\w.\-]+\.\w{2,}", re.I)
_RE_PHONE    = re.compile(r"0[5-7]\d{2}[\s.\-]?\d{2}[\s.\-]?\d{2}[\s.\-]?\d{2}")
_RE_CAR      = re.compile(r"\b(" + "|".join(KNOWN_CARS) + r")\b", re.I)
_RE_COLOR    = re.compile(r"\b(" + "|".join(KNOWN_COLORS) + r")\b", re.I)
_RE_MRZ_NAME = re.compile(r"^P<D[Z7]A([A-Z<]{6,})", re.MULTILINE)   # 容忍 Z/7 OCR 混淆
_RE_PNO_TEXT = re.compile(r"Passport\s*No[^\n]{0,30}\n\s*(\d{7,9})", re.I)  # "Passport No." 下一行数字
_RE_PNO_MRZ2 = re.compile(r"^(\d{9})\d[A-Z]{3}\d{6}", re.MULTILINE)  # MRZ第二行：9位护照号+校验+国籍+生日
_RE_PNO_SOLO = re.compile(r"(?<!\d)(\d{9})(?!\d)")                     # 独立9位数字（最终兜底）
# 注：_RE_PNO_DZA 已删除，DZA后跟数字会误匹配地址

_ADDR_KW = {"cite", "cité", "cete", "céte", "lot", "lgts", "lgmt", "logmt",
            "logmnt", "bt", "uv", "constantine", "oran", "alger", "ain",
            "bey", "mendjeli", "mendjli", "elbey", "adresse", "numero"}


def _extract_from_text(all_text: str, blocks: list[dict], fields: dict):
    """从文字内容中提取各字段，填入 fields（不覆盖已有值）"""

    # ── 邮箱 ──────────────────────────────────────────────
    if not fields["email"]:
        m = _RE_EMAIL.search(all_text)
        if m:
            fields["email"] = m.group().lower().strip(".")

    # ── 电话 ──────────────────────────────────────────────
    if not fields["phone"]:
        m = _RE_PHONE.search(all_text)
        if m:
            fields["phone"] = re.sub(r"[\s.\-]", "", m.group())

    # ── 地址：支持多行合并（如 "...MENDJELI\nCONSTANTINE"） ──
    if not fields["address"]:
        lines = all_text.splitlines()
        for i, line in enumerate(lines):
            tokens = set(re.sub(r"[^a-zA-Z\s]", " ", line).lower().split())
            if tokens & _ADDR_KW and len(line) > 12:
                if _RE_EMAIL.search(line) or "email" in line.lower():
                    continue
                addr_parts = [line.strip()]
                # 检查下一行是否是地址延续
                for j in range(i+1, min(i+3, len(lines))):
                    nxt = lines[j].strip()
                    if not nxt or len(nxt) > 40:
                        break
                    # 遇到电话/邮箱立即停止
                    if _RE_EMAIL.search(nxt) or _RE_PHONE.search(nxt):
                        break
                    nxt_tokens = set(re.sub(r"[^a-zA-Z\s]", " ", nxt).lower().split())
                    # 下一行：含地址关键词 或 纯大写城市名（3-20字母）
                    if nxt_tokens & _ADDR_KW or re.match(r"^[A-Z][A-Za-z]{2,20}$", nxt):
                        addr_parts.append(nxt)
                    else:
                        break
                fields["address"] = " ".join(addr_parts)
                break

    # ── 车型 + 颜色 ────────────────────────────────────────
    if not fields["car_model"]:
        for b in reversed(blocks):
            if b["block_label"] in ("doc_title", "paragraph_title"):
                content = b["block_content"].strip()
                if _RE_CAR.search(content):
                    fields["car_model"] = content
                    m = _RE_COLOR.search(content)
                    if m:
                        fields["color"] = m.group().lower()
                    break
        if not fields["car_model"]:
            for line in reversed(all_text.splitlines()):
                if _RE_CAR.search(line):
                    fields["car_model"] = line.strip()
                    m = _RE_COLOR.search(line)
                    if m:
                        fields["color"] = m.group().lower()
                    break

    # ── 姓名：MRZ 第一行 + Surname 文字区交叉比对 ──────────
    from difflib import SequenceMatcher
    mrz_name = text_surname = None

    for b in blocks:
        mrz_m = _RE_MRZ_NAME.search(b["block_content"])
        if mrz_m:
            raw = mrz_m.group(1).split(" ")[0]
            parts = [p for p in raw.split("<") if p and re.match(r"^[A-Z]{2,}$", p)]
            if len(parts) >= 2:
                mrz_name = f"{' '.join(parts[1:])} {parts[0]}".title()
            elif len(parts) == 1:
                mrz_name = parts[0].title()
            break

    lines = all_text.splitlines()
    for i, line in enumerate(lines):
        u = line.upper()
        if "SURNAME" in u or ("NOM" in u and "/" in u):
            for c in lines[i+1:i+5]:
                c = c.strip()
                if re.match(r"^[A-Z]{2,}(?:\s[A-Z]+)*$", c):
                    text_surname = c.title()
                    break
            break

    if not fields["name"]:
        if mrz_name and text_surname:
            # 文字区通常只有姓（Surname），检查 MRZ 全名是否包含该姓
            surname_in_mrz = text_surname.lower() in mrz_name.lower()
            # 或逐词相似度：文字区每个词在 MRZ 里都能找到相近词
            mrz_words = mrz_name.lower().split()
            text_words = text_surname.lower().split()
            words_match = all(
                any(SequenceMatcher(None, tw, mw).ratio() >= 0.80 for mw in mrz_words)
                for tw in text_words
            )
            if surname_in_mrz or words_match:
                fields["name"] = mrz_name   # MRZ 含全名，更权威
            else:
                # 真正冲突（如 LACHHEB vs LECHHEB）
                fields["name"] = mrz_name
                fields["_name_conflict"] = f"MRZ:{mrz_name} | 文字区:{text_surname}"
        elif mrz_name:
            fields["name"] = mrz_name
        elif text_surname:
            fields["name"] = text_surname

    # ── 护照号：三种来源 ────────────────────────────────────
    if not fields["passport_no"]:
        # 1. "Passport No." 下一行的纯数字（最精确）
        m = _RE_PNO_TEXT.search(all_text)
        if m:
            fields["passport_no"] = m.group(1)

    if not fields["passport_no"]:
        # 2. MRZ 第二行结构：9位护照号 + 校验位 + 国籍码 + 生日
        m = _RE_PNO_MRZ2.search(all_text)
        if m:
            fields["passport_no"] = m.group(1)

    if not fields["passport_no"]:
        # 3. MRZ 第二行兜底：30+位字母数字长串，取前9位，要求含≥4位数字
        for b in blocks:
            for line in b["block_content"].splitlines():
                clean = re.sub(r"[^A-Z0-9]", "", line.upper())
                if len(clean) >= 30 and len(re.findall(r"\d", clean[:9])) >= 4:
                    fields["passport_no"] = clean[:9]
                    break
            if fields["passport_no"]:
                break

    if not fields["passport_no"]:
        # 4. 最终兜底：全文找独立9位纯数字，排除电话号码开头（0[5-7]）
        for m in _RE_PNO_SOLO.finditer(all_text):
            candidate = m.group(1)
            if not re.match(r"^0[5-7]", candidate):  # 排除手机号片段
                fields["passport_no"] = candidate
                break


def extract_fields(full_blocks: list[dict], top_blocks: list[dict]) -> dict:
    """
    两轮提取：
      第一轮用全图 blocks（联系信息、车型、大部分护照数据）
      第二轮用上半裁剪 blocks（补充护照号和姓名）
    """
    fields = {
        "name": None, "passport_no": None,
        "email": None, "phone": None, "address": None,
        "car_model": None, "color": None,
    }

    # 第一轮：全图
    full_text = "\n".join(b["block_content"].strip() for b in full_blocks)
    _extract_from_text(full_text, full_blocks, fields)

    # 第二轮：裁剪上半（仅补充空字段）
    if not fields["name"] or not fields["passport_no"]:
        top_text = "\n".join(b["block_content"].strip() for b in top_blocks)
        _extract_from_text(top_text, top_blocks, fields)

    return fields


# ============================================================
# 字段格式校验（硬规则）
# ============================================================

def validate_fields(fields: dict) -> dict:
    """
    对每个字段做格式校验。
    返回 {field: {"value": ..., "ok": True/False, "reason": ...}}
    """
    results = {}

    def check(field, value, ok, reason=""):
        results[field] = {"value": value, "ok": ok, "reason": reason}

    # 邮箱：标准格式
    email = fields.get("email")
    if email and re.match(r"^[\w.\-+]+@[\w.\-]+\.\w{2,}$", email):
        check("email", email, True)
    else:
        check("email", email, False, "格式异常或未识别" if not email else "邮箱格式不符")

    # 电话：阿尔及利亚手机 05/06/07 开头，10位纯数字
    phone = fields.get("phone")
    if phone and re.match(r"^0[5-7]\d{8}$", phone):
        check("phone", phone, True)
    else:
        check("phone", phone, False, "未识别" if not phone else "不符合阿尔及利亚号码格式(10位/05-07开头)")

    # 地址：非空且长度合理
    addr = fields.get("address")
    if addr and len(addr) >= 10:
        check("address", addr, True)
    else:
        check("address", addr, False, "未识别" if not addr else "地址过短，可能识别不完整")

    # 车型：含已知车型关键词
    car = fields.get("car_model")
    if car and _RE_CAR.search(car):
        check("car_model", car, True)
    else:
        check("car_model", car, False, "未识别" if not car else "未匹配到已知车型关键词")

    # 颜色：在已知颜色列表中
    color = fields.get("color")
    if color and color.lower() in KNOWN_COLORS:
        check("color", color, True)
    else:
        check("color", color, False, "未识别" if not color else "未匹配到已知颜色")

    # 姓名：非空 + 每个词 ≥ 2 字母 + 不含护照标签误识别词
    _NAME_BLACKLIST = {"sport", "port", "sseport", "passeport", "passport",
                       "type", "code", "dza", "nom", "prenom", "surname"}
    name = fields.get("name")
    conflict = fields.get("_name_conflict")
    if not name:
        check("name", name, False, "未识别")
    elif conflict:
        check("name", name, False, f"MRZ与文字区不一致，请人工核对 → {conflict}")
    else:
        words = name.lower().split()
        bad = any(w in _NAME_BLACKLIST or len(w) < 2 for w in words)
        check("name", name, not bad,
              "" if not bad else f"疑似误识别（含异常词: {[w for w in words if w in _NAME_BLACKLIST or len(w)<2]}）")

    # 护照号：阿尔及利亚护照为9位纯数字
    pno = fields.get("passport_no")
    if pno and re.match(r"^\d{9}$", pno):
        check("passport_no", pno, True)
    else:
        check("passport_no", pno, False,
              "未识别" if not pno else f"应为9位纯数字，实际得到: {pno!r}")

    return results


# ============================================================
# 写入 Excel
# ============================================================

_GREEN      = "C6EFCE"
_RED_LIGHT  = "FFE0E0"
_ORANGE     = "FFEB9C"
_HEADER_BG  = "2E75B6"
_WHITE      = "FFFFFF"
_GRAY       = "F2F2F2"

def _hdr_style(cell, bg=_HEADER_BG):
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(bold=True, color=_WHITE, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _row_fill(ws, row_num, color):
    fill = PatternFill("solid", fgColor=color)
    for cell in ws[row_num]:
        cell.fill   = fill
        cell.border = _border()
        cell.alignment = Alignment(vertical="center", wrap_text=True)

def write_excel(all_results: list, output_path: str):
    wb = openpyxl.Workbook()

    # ── Sheet 1：汇总表 ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "收货人汇总"
    ws1.row_dimensions[1].height = 30

    headers = ["文件名", "姓名", "护照号", "邮箱", "电话", "地址", "车型/配置", "颜色", "状态", "待核查字段"]
    ws1.append(headers)
    for i, cell in enumerate(ws1[1], 1):
        _hdr_style(cell)

    # ── Sheet 2：核查详情 ────────────────────────────────
    ws2 = wb.create_sheet("⚠️ 人工核查详情")
    ws2.row_dimensions[1].height = 30
    ws2.append(["文件名", "字段", "识别值", "校验结果", "原因"])
    for cell in ws2[1]:
        _hdr_style(cell, "C00000")

    for r in all_results:
        fname    = r["file"]
        vd       = r["validation"]   # {field: {value, ok, reason}}
        all_ok   = all(v["ok"] for v in vd.values())
        bad_fields = [f for f, v in vd.items() if not v["ok"]]

        def val(f):
            return vd[f]["value"] or ""

        status = "✅ 通过" if all_ok else f"⚠️ {len(bad_fields)} 项待核查"
        row = [
            fname,
            val("name"),
            val("passport_no"),
            val("email"),
            val("phone"),
            val("address"),
            val("car_model"),
            val("color"),
            status,
            "、".join(bad_fields) if bad_fields else "",
        ]
        ws1.append(row)
        row_color = _GREEN if all_ok else _RED_LIGHT
        _row_fill(ws1, ws1.max_row, row_color)

        # Sheet2：写入每个未通过字段的详情
        for field, info in vd.items():
            if not info["ok"]:
                ws2.append([fname, field, info["value"] or "（空）", "❌ 未通过", info["reason"]])
                _row_fill(ws2, ws2.max_row, _RED_LIGHT)

    # 列宽自适应
    col_widths = [18, 20, 14, 30, 14, 45, 28, 10, 14, 24]
    for ws in [ws1, ws2]:
        for i, col in enumerate(ws.columns):
            letter = col[0].column_letter
            w = col_widths[i] if i < len(col_widths) else 18
            ws.column_dimensions[letter].width = w
        ws.freeze_panes = "A2"

    wb.save(output_path)


# ============================================================
# 主流程
# ============================================================

def main():
    pdf_folder = Path(PDF_FOLDER)
    pdf_files  = sorted(pdf_folder.glob("*.pdf"))

    if not pdf_files:
        print("❌ 未找到 PDF 文件，请检查 PDF_FOLDER 路径")
        return

    print(f"📂 找到 {len(pdf_files)} 个 PDF 文件\n")

    all_results = []
    for i, pdf_path in enumerate(pdf_files, 1):
        print(f"[{i}/{len(pdf_files)}] {pdf_path.name} ...", end=" ", flush=True)
        try:
            full_blocks, top_blocks = call_baidu_api(str(pdf_path))
            fields     = extract_fields(full_blocks, top_blocks)
            validation = validate_fields(fields)

            bad = [f for f, v in validation.items() if not v["ok"]]
            if bad:
                print(f"⚠️  待核查: {bad}")
            else:
                print("✅ 全部通过")

            all_results.append({
                "file":       pdf_path.name,
                "fields":     fields,
                "validation": validation,
            })
            time.sleep(0.5)  # 避免触发限流

        except Exception as e:
            print(f"❌ 失败: {e}")
            dummy_validation = {
                f: {"value": None, "ok": False, "reason": f"处理异常: {e}"}
                for f in ["name","passport_no","email","phone","address","car_model","color"]
            }
            all_results.append({"file": pdf_path.name, "fields": {}, "validation": dummy_validation})

    # ── 跨文件交叉校正 ────────────────────────────────────
    cross_validate_across_files(all_results)

    write_excel(all_results, OUTPUT_EXCEL)

    total    = len(all_results)
    all_good = sum(1 for r in all_results if all(v["ok"] for v in r["validation"].values()))
    print(f"\n{'='*50}")
    print(f"📊 总计: {total}  ✅ 全部通过: {all_good}  ⚠️ 需核查: {total - all_good}")
    print(f"📁 Excel: {OUTPUT_EXCEL}")


if __name__ == "__main__":
    main()
